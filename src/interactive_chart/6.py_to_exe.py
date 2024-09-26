from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import os
import sys

# Change application_path to be your current working directory
application_path = os.getcwd()  # This sets it to your script's working directory

month = input('Introduce month: ')

input_path = os.path.join(application_path, 'pivot_table.xlsx')

# Add error handling for file not found
if not os.path.exists(input_path):
    print(f"Error: The file {input_path} was not found.")
    sys.exit(1)

wb = load_workbook(input_path)
sheet = wb['Report']

min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

barchart = BarChart()

data = Reference(sheet,
                 min_col=min_column+1,
                 max_col=max_column,
                 min_row=min_row,
                 max_row=max_row)  # including headers
categories = Reference(sheet,
                       min_col=min_column,
                       max_col=min_column,
                       min_row=min_row+1,
                       max_row=max_row)  # not including headers

barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

sheet.add_chart(barchart, "B12")
barchart.title = 'Sales by Product line'
barchart.style = 5  # choose the chart style

for i in range(min_column+1, max_column+1):  # (B, G+1)
    letter = get_column_letter(i)
    sheet[f'{letter}{max_row + 1}'] = f'=SUM({letter}{min_row + 1}:{letter}{max_row})'
    sheet[f'{letter}{max_row + 1}'].style = 'Currency'

# Format title and month
sheet['A1'] = 'Sales Report'
sheet['A2'] = month
sheet['A1'].font = Font('Arial', bold=True, size=20)
sheet['A2'].font = Font('Arial', bold=True, size=10)

# Save output file to working directory with error handling
output_path = os.path.join(application_path, f'report_{month}.xlsx')

try:
    wb.save(output_path)
    print(f"Report successfully saved at {output_path}")
except Exception as e:
    print(f"Failed to save the report: {e}")
